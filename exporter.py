import openpyxl
from openpyxl import Workbook
from db_worker import db_con
from collections import defaultdict

@db_con
def to_exel(cnn, cur, quiz_id):
    # CREATING WORKBOOK
    x=1
    y=1
    wb=Workbook()

    # USING FIRST SHEET
    ws = wb.active
    ws.title = "Топ сотрудников"

    # WRITING COLUMN NAMES
    columns = ["№ пп", "Участник", "Вопрос", "Результат"]
    for column in columns:
        ws.cell(row=y, column=x).value = column
        x+=1
    y+=1
    x=1

    # GETTING DATA
    cur.execute(f"select user_name, result, try_id from results where quiz_id={quiz_id}")
    results = cur.fetchall()

    res_dict = {}
    for result in results:
        cur.execute(f"select option_id, correct_id from answers where try_id='{result[2]}'")
        q_reses = cur.fetchall()
        print(result)
        print(q_reses)
        q_cnt=1
        res_dict[result[0]] = [result[1], {}]
        for q_res in q_reses:
            if q_res[0]==q_res[1]:
                res_dict[result[0]][1][f"Вопрос {q_cnt}"]=1
            else:
                res_dict[result[0]][1][f"Вопрос {q_cnt}"]=0
            q_cnt+=1

    sorted_dict = sorted(res_dict, key=lambda item: res_dict[item][0], reverse=True)
    print(res_dict)
    print(sorted_dict)
    new_dict = {}
    for item in sorted_dict:
        new_dict[item] = res_dict[item]
    res_dict = new_dict


    # WRITING DATA
    participant_id = 1
    for value in res_dict:
        ws.cell(row=y, column=x).value = participant_id
        x+=1
        ws.cell(row=y, column=x).value = value
        x+=2
        ws.cell(row=y, column=x).value = res_dict[value][0]
        x-=1
        y+=1
        for question in res_dict[value][1]:
            ws.cell(row=y, column=x).value = question
            x+=1
            ws.cell(row=y, column=x).value = res_dict[value][1][question]
            x-=1
            y+=1
        x-=2
        participant_id+=1

    # CREATING NEW SHEET
    ws = wb.create_sheet("Топ неправильных ответов")
    x=1
    y=1

    # WRITING COLUMN NAMES
    columns = ["№ пп", "Участник", "Тест", "Результат"]
    for column in columns:
        ws.cell(row=y, column=x).value = column
        x+=1
    y+=1
    x=1

    # GETTING DATA
    # REUSING DATA FROM PREVIOUS PAGE

    # WRITING DATA
    q_id = 1
    print(f"RES DICT: {res_dict}")
    print(list(res_dict.keys()))
    # might have to pull questions from db, but for now is ok
    question_list = res_dict[list(res_dict.keys())[0]][1]
    incorrect_answers = {}
    for value in question_list:
        incorrect_answers[value] = 0
        for participant in res_dict:
            print(value, participant)
            if res_dict[participant][1][value] == 0:
                incorrect_answers[value] +=1

    print(incorrect_answers, sorted(incorrect_answers, key=lambda item: incorrect_answers[item], reverse=True))

    question_list = sorted(incorrect_answers, key=lambda item: incorrect_answers[item], reverse=True)

    for value in question_list:
        print(f"VALUE: {value}")
        ws.cell(row=y, column=x).value = q_id
        x+=1
        ws.cell(row=y, column=x).value = f"Вопрос {value}"
        x+=2
        total = x, y
        x-=1
        y+=1
        sum_of_incorrect = 0
        for participant in res_dict:
            ws.cell(row=y, column=x).value = participant
            x+=1
            if res_dict[participant][1][value] == 1:
                ws.cell(row=y, column=x).value = 0
            elif res_dict[participant][1][value] == 0:
                ws.cell(row=y, column=x).value = 1
                sum_of_incorrect+=1
            x-=1
            y+=1
        ws.cell(row=total[1], column=total[0]).value = sum_of_incorrect
        x-=2
        q_id+=1

    # CREATING NEW SHEET
    ws = wb.create_sheet("Общий зачёт")
    x=1
    y=1

    # WRITING COLUMN NAMES
    columns = ["№ пп", "Участник", "Тест", "Результат"]
    for column in columns:
        ws.cell(row=y, column=x).value = column
        x+=1
    y+=1
    x=1

    # GETTING DATA
    cur.execute(f"select user_name, quiz_id, result from results")
    results = cur.fetchall()
    print(results)
    grouped_dict = defaultdict(list)
    for u, q, r in results: 
        grouped_dict[u].append((q,r))
    print(grouped_dict)

    # WRITING DATA
    participant_id = 1
    for value in grouped_dict:
        ws.cell(row=y, column=x).value = participant_id
        x+=1
        ws.cell(row=y, column=x).value = value
        x+=2
        total = x, y
        x-=1
        y+=1
        sum_of_points = 0
        for item in grouped_dict[value]:
            q, r = item
            ws.cell(row=y, column=x).value = q
            x+=1
            ws.cell(row=y, column=x).value = r
            sum_of_points+=int(r)
            x-=1
            y+=1
        x-=2
        ws.cell(row=total[1], column=total[0]).value = sum_of_points
        participant_id+=1

    # SAVING
    wb.save(f"./export/{quiz_id}.xlsx")
    
if __name__ == "__main__":
    to_exel(37)